import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_changes(history_file):
    # Load the data using pandas to easily find CN columns
    df = pd.read_excel(history_file)
    cn_cols = [col for col in df.columns if 'CN' in col]
    
    # Load the workbook for styling
    wb = load_workbook(history_file)
    ws = wb.active
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Map column names to 1-based indexing for openpyxl
    header = [cell.value for cell in ws[1]]
    col_map = {name: i+1 for i, name in enumerate(header)}
    
    # Iterate through rows (skipping header)
    for row_idx in range(2, ws.max_row + 1):
        # We compare each CN column to the one directly preceding it in our sorted list
        for i in range(1, len(cn_cols)):
            prev_col_name = cn_cols[i-1]
            curr_col_name = cn_cols[i]
            
            prev_val = ws.cell(row=row_idx, column=col_map[prev_col_name]).value
            curr_val = ws.cell(row=row_idx, column=col_map[curr_col_name]).value
            
            # If current value exists and is different from previous (not just appearing for the first time)
            if curr_val and prev_val and curr_val != prev_val:
                ws.cell(row=row_idx, column=col_map[curr_col_name]).fill = yellow_fill
            
            # Special case: handle the "Current Text CN" if it's new/different
            # (Already covered if Current Text CN is in cn_cols and we iterated through it)

    wb.save(history_file)
    print(f"Highlighted changes in {history_file}")

if __name__ == "__main__":
    work_dir = '/Users/gardsale/Projects/Makeup - YELLOW - 2/'
    history_file = os.path.join(work_dir, 'text_history.xlsx')
    if os.path.exists(history_file):
        highlight_changes(history_file)
    else:
        print("History file not found.")
